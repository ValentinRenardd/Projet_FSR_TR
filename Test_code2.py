# Ne pas oublier d'installer pyserial et serial et autres si besoin...
import numpy as np
import tkinter as tk
import serial
import time
import matplotlib.pyplot as plt
import pandas as pd
import threading
from tkinter import messagebox
from openpyxl import Workbook


# Configuration du port série
baudrate = 9600
duration = 10  # Durée d'acquisition en secondes


# Fonction pour démarrer le décompte et l'enregistrement simultanément
def start_acquisition():
    submit_button.pack_forget()  # Cache le bouton "Démarrer"
    # Démarre le décompte dans un thread séparé
    threading.Thread(target=start_countdown, args=(duration,)).start()
    # Démarre l'enregistrement des données
    threading.Thread(target=record_data).start()

# Fonction pour démarrer le décompte
def start_countdown(time_left):
    while time_left >= 0:
        # Met à jour le label avec le temps restant
        time_label.config(text=f"Temps restant : {time_left} secondes")
        # Attendre 1 seconde avant de décrémenter le temps
        time.sleep(1)
        time_left -= 1
    
    # Lorsque le temps est écoulé, afficher un message
    time_label.config(text="Le temps est écoulé!")

# Fonction pour enregistrer les données
def record_data():
    data = []
    time1 =[]
    
    nom = entry_nom.get().strip()
    prenom = entry_prenom.get().strip()
    essai = entry_essai.get().strip()
    port_com = entry_port_com.get().strip()
    
    if not nom or not prenom or not essai or not port_com:
        messagebox.showerror("Erreur", "Tous les champs sont obligatoires !")
        return
    
    # Initialisation du port série
    ser = serial.Serial(port_com, baudrate, timeout=1)
    print("Acquisition en cours...")

    # Enregistrement des données
    start_time = time.time()
    
    try:
        while time.time() - start_time < duration:  # Assure une durée de 10 secondes
            line = ser.readline().decode('utf-8').strip()  # Lecture d'une ligne
            if line.isdigit():  # Vérifie si la ligne contient une valeur numérique
                data.append(int(line))  # Stocke la valeur
                time1.append(time.time())  # Stocke la valeur
    except KeyboardInterrupt:
        print("Acquisition interrompue.")
    
    # Fermeture du port série
    ser.close()
    
    # Appeler la fonction pour compter les pics
    peaks_count = count_peaks(data, seuil_min=15)  # Distance minimale = 50 indices
    

    
    fichier_nom = f"{nom}_{prenom}_Essai{essai}.xlsx"
    vec_temps = [x - time1[0] for x in time1]
    
    try:
        # Création du fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Données Acquises"

        
        ws.append(["Temps Machine (s)", "Temps Acquisition (s)", "Valeur du capteur", "Nombre de clics"])
        
        # Ajouter les vecteurs comme colonnes
        for i, (t, vec_t, val) in enumerate(zip(time1, vec_temps, data)):
            # Ajouter les éléments de time1 (Temps Machine) dans la colonne 1
            ws.cell(row=i+2, column=1, value=t)  # Ligne i+2 pour commencer à partir de la ligne 2
            # Ajouter les éléments de vec_temps (Temps Acquisition) dans la colonne 2
            ws.cell(row=i+2, column=2, value=vec_t)
            # Ajouter les éléments de data (Valeur du capteur) dans la colonne 3
            ws.cell(row=i+2, column=3, value=val)
        
        # Ajouter le nombre de clics dans la 4e colonne, ligne 2 seulement
        ws.cell(row=2, column=4, value=peaks_count)  # Ligne 2, colonne 4

        # Sauvegarde du fichier
        wb.save(fichier_nom)
        messagebox.showinfo("Succès", f"Le fichier '{fichier_nom}' a été créé avec les données acquises et traitées.")
    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")


    # Afficher les données dans un graphique
    plt.plot(vec_temps, data)
    plt.xlabel("Temps (s)")
    plt.ylabel("Valeur du capteur")
    plt.title("Données FSR")
    plt.show()  # Affiche le graphique

    
    
    # Afficher le nombre de pics trouvés
    print(f"Nombre de pics trouvés : {peaks_count}")
    messagebox.showinfo("Résultat", f"Nombre de clics :{peaks_count}")

# Fonction pour compter le nombre de pics au-dessus d'un seuil donné
# Ajout d'un paramètre min_distance pour fusionner les pics proches
def count_peaks(data, seuil_min=15):
    peaks_count = 0

    for i in range(1, len(data) - 1):
        # Condition pour un pic (plus grand que ses voisins)
        if data[i] > data[i - 1] and data[i] > seuil_min and data[i-1] < seuil_min:
            peaks_count += 1  # Un nouveau pic est trouvé

    return peaks_count

    
# Création de la fenêtre principale
root = tk.Tk()
root.title("Formulaire de Participant")

# Nom
tk.Label(root, text="Nom :").grid(row=0, column=0, padx=10, pady=5, sticky="e")
entry_nom = tk.Entry(root)
entry_nom.grid(row=0, column=1, padx=10, pady=5)

# Prénom
tk.Label(root, text="Prénom :").grid(row=1, column=0, padx=10, pady=5, sticky="e")
entry_prenom = tk.Entry(root)
entry_prenom.grid(row=1, column=1, padx=10, pady=5)

# Numéro d'essai
tk.Label(root, text="Numéro d'essai :").grid(row=2, column=0, padx=10, pady=5, sticky="e")
entry_essai = tk.Entry(root)
entry_essai.grid(row=2, column=1, padx=10, pady=5)

# Port COM
tk.Label(root, text="Port COM :").grid(row=3, column=0, padx=10, pady=5, sticky="e")
entry_port_com = tk.Entry(root)
entry_port_com.grid(row=3, column=1, padx=10, pady=5)

# Bouton Soumettre
submit_button = tk.Button(root, text="Démarrer", command=start_acquisition)
submit_button.grid(row=4, column=0, columnspan=2, pady=10)


# Création du label pour afficher le décompte
time_label = tk.Label(root, text=f"Temps restant : {duration} secondes", font=("Arial", 16))
time_label.grid(row=5, column=0, columnspan=2, pady=20)



# Lancement de la boucle principale
root.mainloop()
